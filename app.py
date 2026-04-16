# ============================================================
#  app.py — IT Procurement Intelligence Dashboard
#  Dash by Plotly | PwC Brand | Source Sans Pro
# ============================================================

import dash
from dash import dcc, html, Input, Output, State, ctx, dash_table
import dash_bootstrap_components as dbc
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import numpy as np
import base64
import io
import os
import re
import zipfile
import difflib
from collections import defaultdict
import openpyxl

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    import requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

# ════════════════════════════════════════════════════════════
# INIT APP
# ════════════════════════════════════════════════════════════
app = dash.Dash(
    __name__,
    external_stylesheets=[
        dbc.themes.BOOTSTRAP,
        "https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@300;400;600;700&display=swap",
    ],
    suppress_callback_exceptions=True,
    title="IT Procurement Intelligence",
)
server = app.server  # for deployment

# ════════════════════════════════════════════════════════════
# PwC COLOURS
# ════════════════════════════════════════════════════════════
PWC = {
    "orange" : "#D04A02",
    "dark"   : "#2D2D2D",
    "blue"   : "#295477",
    "teal"   : "#299D8F",
    "yellow" : "#FFB600",
    "green"  : "#22992E",
    "red"    : "#E0301E",
    "purple" : "#6E2585",
    "grey"   : "#8C8C8C",
    "light"  : "#F3F3F3",
    "white"  : "#FFFFFF",
}
PALETTE = [
    "#D04A02","#295477","#299D8F",
    "#FFB600","#22992E","#E0301E",
    "#EB8C00","#6E2585","#8C8C8C","#004F9F",
]
def get_color(i):
    return PALETTE[i % len(PALETTE)]

CFONT = dict(
    family="Source Sans Pro, Helvetica Neue, Arial",
    size=11, color="#2D2D2D")
CBG = "#F3F3F3"

# ════════════════════════════════════════════════════════════
# GLOBAL CSS
# ════════════════════════════════════════════════════════════
CUSTOM_CSS = """
<style>
* { font-family: 'Source Sans Pro','Helvetica Neue',Arial,sans-serif !important; }
body { background:#F3F3F3 !important; }

/* Sidebar */
.sidebar {
    background:#2D2D2D !important;
    border-right:3px solid #D04A02;
    min-height:100vh;
    padding:0 0 20px;
}

/* Nav tabs */
.nav-tabs .nav-link {
    color:#7D7D7D !important;
    font-weight:600 !important;
    border:none !important;
    border-bottom:3px solid transparent !important;
    padding:10px 16px !important;
}
.nav-tabs .nav-link.active {
    color:#D04A02 !important;
    border-bottom:3px solid #D04A02 !important;
    background:transparent !important;
}
.nav-tabs { border-bottom:2px solid #ddd !important; }

/* KPI */
.kpi-card {
    border-radius:4px;
    padding:18px 12px;
    text-align:center;
    color:white;
    border-left:5px solid rgba(255,255,255,0.25);
}
.kpi-value {
    font-size:2.1em;
    font-weight:700;
    line-height:1.1;
}
.kpi-label {
    font-size:0.76em;
    font-weight:700;
    opacity:0.9;
    text-transform:uppercase;
    letter-spacing:0.8px;
    margin-top:4px;
}

/* Tables */
.pwc-table { width:100%; border-collapse:collapse; font-size:0.83em; table-layout:fixed; }
.pwc-table thead tr { background:#2D2D2D; }
.pwc-table thead th {
    padding:10px 10px; text-align:left;
    font-weight:700; font-size:0.80em;
    letter-spacing:0.4px; text-transform:uppercase;
    color:white !important; border:none; word-break:break-word;
}
.pwc-table tbody tr:nth-child(even) { background:#F3F3F3; }
.pwc-table tbody tr:hover { background:#FCE8DC; }
.pwc-table tbody td {
    padding:8px 10px; border-bottom:1px solid #e8e8e8;
    vertical-align:middle; word-break:break-word;
    font-size:0.82em; color:#2D2D2D;
}

/* Vendor badge */
.vbadge {
    display:inline-block; padding:3px 8px;
    border-radius:2px; color:white;
    font-size:0.78em; font-weight:700;
    white-space:nowrap; overflow:hidden;
    text-overflow:ellipsis; max-width:120px;
}

/* Score cards */
.score-card {
    border-radius:4px; padding:14px 16px;
    margin-bottom:10px; border-left:5px solid #D04A02;
}
.score-card.green  { background:#F0FFF4; border-color:#22992E; }
.score-card.yellow { background:#FFF8E1; border-color:#FFB600; }
.score-card.red    { background:#FFF3F0; border-color:#E0301E; }

/* AI box */
.ai-box {
    background:#F8F0FF;
    border-left:5px solid #6E2585;
    border-radius:4px;
    padding:14px 18px;
    margin:10px 0;
}

/* Expander fix */
.accordion-button:not(.collapsed) {
    color:#D04A02 !important;
    background:white !important;
    box-shadow:none !important;
}
.accordion-button::after {
    display:none !important;
}
.accordion-button {
    font-weight:700 !important;
    font-size:0.92em !important;
}

/* Upload zone */
.upload-zone {
    border:2px dashed #D04A02;
    border-radius:8px;
    padding:30px;
    text-align:center;
    background:white;
    cursor:pointer;
}

/* Section title */
.section-title {
    font-size:0.72em;
    font-weight:700;
    letter-spacing:1px;
    text-transform:uppercase;
    color:#D04A02;
    margin-bottom:4px;
}

/* Sidebar labels */
.sb-label {
    color:#F0F0F0;
    font-weight:700;
    font-size:0.82em;
    letter-spacing:0.5px;
    text-transform:uppercase;
    padding:10px 16px 4px;
}

/* Step box */
.step-box {
    background:white;
    border-left:4px solid #D04A02;
    border-radius:4px;
    padding:14px 18px;
    margin-bottom:12px;
    box-shadow:0 1px 4px rgba(0,0,0,0.06);
}
.step-num {
    font-size:0.68em;
    font-weight:700;
    letter-spacing:1px;
    text-transform:uppercase;
    color:#D04A02;
    margin-bottom:4px;
}
</style>
"""

# ════════════════════════════════════════════════════════════
# PRICE EXTRACTION
# ════════════════════════════════════════════════════════════
PRICE_RE = re.compile(
    r"""
    (?:USD|EUR|GBP|SGD|MYR|AUD|CAD)\s?\d{1,3}(?:[,]\d{3})*(?:\.\d{1,2})?
    |(?:[\$\€\£]\s?)\d{1,3}(?:[,\s]\d{3})*(?:\.\d{1,2})?
    |\d{1,3}(?:[,]\d{3})+(?:\.\d{1,2})?
    """,
    re.VERBOSE | re.IGNORECASE,
)
TOTAL_KW = [
    "grand total","total amount","total price","amount due",
    "net total","total cost","total value","quote total",
    "subtotal","estimated total","total",
]

def _parse_num(s):
    try:
        return float(re.sub(r"[^\d.]","",str(s)) or "0")
    except:
        return 0.0

def _fmt(val):
    try:
        v = float(re.sub(r"[^\d.]","",str(val)) or "0")
        if v <= 0: return "—"
        return "${:,.2f}".format(v)
    except:
        return str(val)

def _best_price(text):
    tl = text.lower()
    for kw in TOTAL_KW:
        idx = tl.find(kw)
        if idx == -1: continue
        snip  = text[max(0,idx-20):idx+300]
        hits  = PRICE_RE.findall(snip)
        valid = [h.strip() for h in hits if _parse_num(h)>=50]
        if valid:
            return max(valid, key=_parse_num)
    all_h = PRICE_RE.findall(text)
    valid = [h.strip() for h in all_h if _parse_num(h)>=100]
    if valid:
        return max(valid, key=_parse_num)
    return ""

def _text_from_bytes(content, ext):
    text = ""
    ext  = ext.lower().strip(".")
    try:
        if ext == "pdf":
            if not PDF_OK: return ""
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for p in pdf.pages:
                    t = p.extract_text()
                    if t: text += t + "\n"
        elif ext in ("xlsx","xls"):
            wb = openpyxl.load_workbook(
                io.BytesIO(content), data_only=True, read_only=True)
            rows_text = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    rs = "  ".join(str(c) for c in row if c is not None)
                    if rs.strip(): rows_text.append(rs)
            text = "\n".join(rows_text)
            wb.close()
        elif ext == "docx":
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                if "word/document.xml" in z.namelist():
                    xml  = z.read("word/document.xml").decode("utf-8",errors="ignore")
                    text = re.sub(r"<[^>]+>"," ",xml)
                    text = re.sub(r"\s{2,}","\n",text)
    except:
        pass
    return text

def extract_price_from_bytes(content, ext):
    text  = _text_from_bytes(content, ext)
    price = _best_price(text)
    return {
        "price"    : price,
        "price_num": _parse_num(price) if price else 0.0,
        "text"     : text[:5000],
    }

def extract_price_from_url(url):
    if not REQUESTS_OK:
        return {"price":"","price_num":0.0,"text":""}
    try:
        r   = requests.get(url, timeout=20)
        ext = url.split("?")[0].rsplit(".",1)[-1].lower()
        return extract_price_from_bytes(r.content, ext)
    except:
        return {"price":"","price_num":0.0,"text":""}

# ════════════════════════════════════════════════════════════
# SCORING
# ════════════════════════════════════════════════════════════
def price_score(new_price, hist_prices):
    valid = [p for p in hist_prices if p > 0]
    if not valid or new_price <= 0:
        return None, "No comparison data", 0, 0, 0
    mn  = min(valid)
    mx  = max(valid)
    avg = sum(valid) / len(valid)
    if mx == mn:
        return 50, "Same as historical average", avg, mn, mx
    score = round((1-(new_price-mn)/(mx-mn))*100, 1)
    score = max(0, min(100, score))
    pct   = round((new_price-avg)/avg*100, 1)
    if new_price < avg:
        label = "{}% BELOW average — COMPETITIVE".format(abs(pct))
    elif new_price > avg:
        label = "{}% ABOVE average — REVIEW NEEDED".format(abs(pct))
    else:
        label = "Matches historical average"
    return score, label, avg, mn, mx

def score_color(s):
    if s is None: return PWC["grey"]
    if s >= 70:   return PWC["green"]
    if s >= 40:   return PWC["yellow"]
    return PWC["red"]

def score_css(s):
    if s is None: return "yellow"
    if s >= 70:   return "green"
    if s >= 40:   return "yellow"
    return "red"

def service_similarity(sa, sb):
    if not sa or not sb: return 0.0
    a = set(s.lower().strip() for s in sa)
    b = set(s.lower().strip() for s in sb)
    if not a or not b: return 0.0
    return round(len(a&b)/len(a|b)*100, 1)

# ════════════════════════════════════════════════════════════
# AI INSIGHTS
# ════════════════════════════════════════════════════════════
def ai_service_summary(df_master, df_exploded):
    svc_by_v = {}
    for v in df_master["Vendor"].unique():
        svc_by_v[v] = list(
            df_exploded[df_exploded["Vendor"]==v]["Service"].unique())
    if not svc_by_v:
        return "No vendor data available."
    best   = max(svc_by_v, key=lambda v: len(svc_by_v[v]))
    n_best = len(svc_by_v[best])
    total  = len(set(s for svcs in svc_by_v.values() for s in svcs))
    shared = [
        s for s in set(
            s for svcs in svc_by_v.values() for s in svcs)
        if sum(1 for svcs in svc_by_v.values() if s in svcs) > 1
    ]
    lines = [
        "**{}** covers the most services ({} of {} total).".format(
            best, n_best, total)
    ]
    if shared:
        lines.append(
            "**{}** service(s) offered by multiple vendors — "
            "ideal for competitive benchmarking.".format(len(shared)))
    return " ".join(lines)

def ai_price_insight(new_price, hist_prices, vendor_prices):
    valid = [p for p in hist_prices if p > 0]
    if not valid or new_price <= 0:
        return "Insufficient data for price analysis."
    avg = sum(valid) / len(valid)
    mn  = min(valid)
    mx  = max(valid)
    pct = round((new_price-avg)/avg*100, 1)
    lines = []
    if new_price <= mn:
        lines.append("This quote is the **lowest price** seen — excellent value.")
    elif new_price >= mx:
        lines.append("This quote is **above all historical prices** — negotiate strongly.")
    elif pct > 15:
        lines.append("Quote is **{}% above** average. Request a revised quote.".format(abs(pct)))
    elif pct < -15:
        lines.append("Quote is **{}% below** average — very competitive.".format(abs(pct)))
    else:
        lines.append("Quote is **within normal range** ({}% vs average).".format(pct))
    if vendor_prices:
        best_v = min(vendor_prices, key=vendor_prices.get)
        lines.append("**{}** has historically offered the lowest prices.".format(best_v))
    return " ".join(lines)

def ai_analyze_catalog(df, df_exp):
    insights = {}
    n_vendors  = df["Vendor"].nunique()
    n_files    = df["File Name"].nunique()
    n_cats     = df["Category"].nunique()
    n_services = df_exp["Service"].nunique() if not df_exp.empty else 0
    insights["overview"] = (
        "Catalog contains **{} vendors**, **{} quote files**, "
        "**{} categories** and **{} unique services**.".format(
            n_vendors, n_files, n_cats, n_services))
    spv = df_exp.groupby("Vendor")["Service"].nunique().sort_values(ascending=False)
    if not spv.empty:
        top_v = spv.index[0]
        insights["top_vendor"] = (
            "**{}** leads with **{}** unique services — "
            "strongest overall coverage.".format(top_v, spv.iloc[0]))
    shared = df_exp.groupby("Service")["Vendor"].nunique().sort_values(ascending=False)
    hot = shared[shared > 1]
    if not hot.empty:
        insights["competitive"] = (
            "**{}** is the most competitive service with "
            "**{}** vendors quoting.".format(hot.index[0], hot.iloc[0]))
    cat_counts = (
        df.drop_duplicates(subset=["Category","File Name"])
        .groupby("Category").size().sort_values(ascending=False))
    if not cat_counts.empty:
        pct = round(cat_counts.iloc[0]/cat_counts.sum()*100, 1)
        insights["category"] = (
            "**{}** dominates with **{}%** of all quote files.".format(
                cat_counts.index[0], pct))
    if "Quoted Price" in df.columns:
        prices = df["Quoted Price"].apply(_parse_num)
        prices = prices[prices > 0]
        if not prices.empty:
            insights["pricing"] = (
                "Quoted prices range from **{}** to **{}** "
                "with average **{}**.".format(
                    _fmt(prices.min()), _fmt(prices.max()),
                    _fmt(prices.mean())))
    recs = []
    if not hot.empty and len(hot) >= 3:
        recs.append("Run competitive bids on {} services offered by multiple vendors.".format(len(hot)))
    if n_vendors >= 3:
        recs.append("Consider vendor consolidation — {} vendors may create management overhead.".format(n_vendors))
    insights["recommendations"] = recs
    return insights

# ════════════════════════════════════════════════════════════
# DATA HELPERS
# ════════════════════════════════════════════════════════════
def group_services(services):
    groups = defaultdict(list)
    for svc in sorted(services):
        tokens = str(svc).strip().split()
        if not tokens:
            groups["Other"].append(svc)
            continue
        key = tokens[0]
        if len(key) <= 3 and len(tokens) > 1:
            key = tokens[0] + " " + tokens[1]
        groups[key.strip()].append(svc)
    final = {}
    other = []
    for k, v in groups.items():
        if len(v) >= 2: final[k] = sorted(v)
        else: other.extend(v)
    if other: final["Other"] = sorted(other)
    return dict(sorted(final.items()))

@staticmethod
def extract_hyperlinks_from_wb(wb):
    link_map = {}
    try:
        ws = wb.active
        fn_col = hr = None
        for row in ws.iter_rows():
            for cell in row:
                if (cell.value and
                        str(cell.value).strip().lower() == "file name"):
                    fn_col = cell.column
                    hr     = cell.row
                    break
            if fn_col: break
        if fn_col:
            for row in ws.iter_rows(
                    min_row=hr+1, min_col=fn_col, max_col=fn_col):
                cell = row[0]
                if cell.value and cell.hyperlink:
                    link_map[str(cell.value).strip()] = \
                        str(cell.hyperlink.target).strip()
    except:
        pass
    return link_map

def process_catalog(file_bytes, filename):
    try:
        ext = filename.rsplit(".",1)[-1].lower()
        if ext in ("xlsx","xls"):
            wb  = openpyxl.load_workbook(
                io.BytesIO(file_bytes))
            hmap = {}
            try:
                ws = wb.active
                fn_col = hr = None
                for row in ws.iter_rows():
                    for cell in row:
                        if (cell.value and
                                str(cell.value).strip().lower()
                                == "file name"):
                            fn_col = cell.column
                            hr     = cell.row
                            break
                    if fn_col: break
                if fn_col:
                    for row in ws.iter_rows(
                            min_row=hr+1,
                            min_col=fn_col,
                            max_col=fn_col):
                        cell = row[0]
                        if cell.value and cell.hyperlink:
                            hmap[str(cell.value).strip()] = \
                                str(cell.hyperlink.target).strip()
            except:
                pass
            raw = pd.read_excel(
                io.BytesIO(file_bytes),
                engine="openpyxl", header=None)
        elif ext == "csv":
            raw  = pd.read_csv(
                io.BytesIO(file_bytes), header=None)
            hmap = {}
        else:
            return None, None, "Unsupported file type."

        # Detect header row
        header_row = 0
        for i, row in raw.iterrows():
            vals = [str(v).strip().lower()
                    for v in row.values if pd.notna(v)]
            joined = " ".join(vals)
            has_v  = any(k in joined
                for k in ["vendor","supplier","company"])
            has_f  = any(k in joined
                for k in ["file","document","attachment"])
            has_c  = any(k in joined
                for k in ["category","type","domain"])
            if (has_v or has_c) and has_f:
                header_row = i
                break

        if ext in ("xlsx","xls"):
            df = pd.read_excel(
                io.BytesIO(file_bytes),
                engine="openpyxl", header=header_row)
        else:
            df = pd.read_csv(
                io.BytesIO(file_bytes), header=header_row)

        df = df.loc[:, df.columns.notna()]
        df.columns = [str(c).strip() for c in df.columns]
        df.dropna(how="all", inplace=True)

        # Column mapping
        col_map = {}
        for c in df.columns:
            cl = str(c).lower().strip()
            if any(k in cl for k in ["category","type","domain"]):
                if "Category" not in col_map: col_map["Category"] = c
            elif any(k in cl for k in ["vendor","supplier","company","provider"]):
                if "Vendor" not in col_map: col_map["Vendor"] = c
            elif any(k in cl for k in ["file name","filename","document","attachment"]):
                if "File Name" not in col_map: col_map["File Name"] = c
            elif any(k in cl for k in ["link","url","hyperlink"]):
                if "File Link" not in col_map: col_map["File Link"] = c
            elif any(k in cl for k in ["comment","service","description","scope","remark"]):
                if "Comments" not in col_map: col_map["Comments"] = c
            elif any(k in cl for k in ["price","cost","amount","value","quote","rate"]):
                if "Quoted Price" not in col_map: col_map["Quoted Price"] = c

        df.rename(columns={v:k for k,v in col_map.items()}, inplace=True)

        for req in ["Category","Vendor","File Name"]:
            if req not in df.columns: df[req] = ""
        if "Comments" not in df.columns: df["Comments"] = ""

        keep = ["Category","Vendor","File Name","Comments"]
        for e in ["File Link","Quoted Price"]:
            if e in df.columns: keep.append(e)
        df = df[[c for c in keep if c in df.columns]].copy()

        df = df[
            ~(df["Category"].astype(str).str.strip().isin(["","nan"]) &
              df["Vendor"].astype(str).str.strip().isin(["","nan"]))
        ].copy()

        for col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()

        df["Hyperlink"] = df["File Name"].map(hmap).fillna("")

        def parse_svc(v):
            if not v or str(v).strip() in ["","nan"]:
                return ["(unspecified)"]
            parts = [s.strip() for s in str(v).split("\n") if s.strip()]
            return parts or ["(unspecified)"]

        df["Services List"] = df["Comments"].apply(parse_svc)
        df.reset_index(drop=True, inplace=True)

        df_exp = df.explode("Services List").copy()
        df_exp.rename(columns={"Services List":"Service"}, inplace=True)
        df_exp["Service"] = df_exp["Service"].str.strip()
        df_exp = df_exp[
            ~df_exp["Service"].isin(["","(unspecified)","nan"])
        ].reset_index(drop=True)

        return df, df_exp, None
    except Exception as e:
        return None, None, str(e)

# ════════════════════════════════════════════════════════════
# LOAD DEFAULT DATA
# ════════════════════════════════════════════════════════════
def load_default():
    FILE_PATH = "Master Catalog.xlsx"
    if not os.path.exists(FILE_PATH):
        return None, None
    with open(FILE_PATH,"rb") as f:
        return process_catalog(f.read(), FILE_PATH)[:2]

df_default, df_exp_default = load_default()
if df_default is None:
    df_default  = pd.DataFrame(columns=[
        "Category","Vendor","File Name",
        "Comments","Hyperlink","Services List"])
    df_exp_default = pd.DataFrame(columns=[
        "Category","Vendor","File Name",
        "Comments","Hyperlink","Service"])

# ════════════════════════════════════════════════════════════
# CHART BUILDERS
# ════════════════════════════════════════════════════════════
def build_overlap_chart(df_exp):
    if df_exp.empty:
        return go.Figure()
    shared = (df_exp.groupby("Service")["Vendor"].nunique()
              .sort_values(ascending=False).head(20).reset_index())
    shared.columns = ["Service","Vendor Count"]
    shared["Color"] = shared["Vendor Count"].apply(
        lambda x: PWC["orange"] if x > 1 else "#C0C0C0")
    fig = go.Figure(go.Bar(
        x=shared["Vendor Count"],
        y=shared["Service"].str[:44],
        orientation="h",
        marker_color=shared["Color"],
        marker_line_width=0,
        text=shared["Vendor Count"],
        textposition="outside",
        textfont=dict(size=10),
    ))
    fig.update_layout(
        height=480, plot_bgcolor=CBG, paper_bgcolor=CBG,
        margin=dict(l=5,r=40,t=20,b=10), font=CFONT,
        xaxis=dict(title="Vendors",showgrid=True,
                   gridcolor="#E0E0E0",zeroline=False),
        yaxis=dict(autorange="reversed",tickfont=dict(size=9.5)),
        bargap=0.35)
    return fig

def build_coverage_chart(df_exp, vcmap):
    if df_exp.empty:
        return go.Figure()
    spv = (df_exp.groupby("Vendor")["Service"].nunique()
           .sort_values(ascending=False).reset_index())
    spv.columns = ["Vendor","Count"]
    spv["Color"] = [vcmap.get(v,PWC["grey"]) for v in spv["Vendor"]]
    fig = go.Figure(go.Bar(
        x=spv["Vendor"], y=spv["Count"],
        marker_color=spv["Color"],
        marker_line_width=0,
        text=spv["Count"],
        textposition="outside",
        textfont=dict(size=10),
    ))
    fig.update_layout(
        height=480, plot_bgcolor=CBG, paper_bgcolor=CBG,
        margin=dict(l=5,r=10,t=20,b=10), font=CFONT,
        yaxis=dict(title="Unique Services",showgrid=True,
                   gridcolor="#E0E0E0",zeroline=False),
        xaxis=dict(tickangle=-35,tickfont=dict(size=9.5)),
        bargap=0.35)
    return fig

def build_category_chart(df_exp):
    if df_exp.empty:
        return go.Figure()
    cat_c = (df_exp.drop_duplicates(subset=["Category","File Name"])
             .groupby("Category").size().reset_index())
    cat_c.columns = ["Category","Count"]
    fig = px.pie(cat_c, names="Category", values="Count",
                 hole=0.50, color_discrete_sequence=PALETTE)
    fig.update_traces(
        textposition="outside", textinfo="label+percent",
        textfont_size=11, pull=[0.03]*len(cat_c))
    fig.update_layout(
        height=380, margin=dict(l=20,r=20,t=20,b=20),
        paper_bgcolor=CBG, font=CFONT,
        legend=dict(orientation="v",x=1.02,y=0.5,
                    font=dict(size=10)))
    return fig

def build_price_chart(price_data, avg_price):
    if not price_data:
        return go.Figure()
    df_p = pd.DataFrame(price_data)
    fig  = go.Figure()
    hist = df_p[df_p["Type"]=="Historical"]
    new  = df_p[df_p["Type"]=="New"]
    if not hist.empty:
        fig.add_trace(go.Bar(
            x=hist["Label"], y=hist["Price"],
            marker_color=hist["Color"],
            marker_line_width=0,
            name="Historical",
            text=hist["Price"].apply(_fmt),
            textposition="outside"))
    if not new.empty:
        fig.add_trace(go.Bar(
            x=new["Label"], y=new["Price"],
            marker_color=PWC["orange"],
            marker_line_width=0,
            name="New Upload",
            text=new["Price"].apply(_fmt),
            textposition="outside"))
    if avg_price > 0:
        fig.add_hline(
            y=avg_price, line_dash="dash",
            line_color=PWC["yellow"], line_width=2,
            annotation_text="Avg: {}".format(_fmt(avg_price)),
            annotation_position="top right")
    fig.update_layout(
        height=360, plot_bgcolor=CBG, paper_bgcolor=CBG,
        margin=dict(l=5,r=10,t=20,b=10), font=CFONT,
        barmode="group",
        yaxis=dict(title="Price",showgrid=True,
                   gridcolor="#E0E0E0",zeroline=False),
        xaxis=dict(tickangle=-25),
        legend=dict(orientation="h",x=0,y=1.05),
        bargap=0.25)
    return fig

# ════════════════════════════════════════════════════════════
# HTML HELPERS
# ════════════════════════════════════════════════════════════
def vendor_pill(v, color):
    return html.Span(
        v,
        className="vbadge",
        style={"background":color,
               "marginRight":"4px"})

def section_title_div(txt, caption=""):
    children = [
        html.Div(txt, className="section-title")]
    if caption:
        children.append(
            html.Div(caption,
                style={"fontSize":"0.80em",
                       "color":"#7D7D7D",
                       "marginBottom":"10px"}))
    return html.Div(children)

def ai_box_div(content):
    return html.Div([
        html.Div("🤖 AI Insight",
            style={"fontSize":"0.68em","fontWeight":"700",
                   "letterSpacing":"1px",
                   "textTransform":"uppercase",
                   "color":PWC["purple"],
                   "marginBottom":"6px"}),
        dcc.Markdown(content,
            style={"fontSize":"0.87em",
                   "color":PWC["dark"]})
    ], className="ai-box")

def score_card_div(label, value, sub, css_class, color):
    return html.Div([
        html.Div(label,
            style={"fontSize":"0.68em","fontWeight":"700",
                   "letterSpacing":"1px",
                   "textTransform":"uppercase",
                   "color":color}),
        html.Div(value,
            style={"fontSize":"2.0em","fontWeight":"800",
                   "lineHeight":"1","color":color}),
        html.Div(sub,
            style={"fontSize":"0.75em",
                   "color":"#555","marginTop":"4px"}),
    ], className="score-card {}".format(css_class))

def kpi_card(value, label, bg):
    return html.Div([
        html.Div(str(value), className="kpi-value"),
        html.Div(label, className="kpi-label"),
    ], className="kpi-card",
       style={"background":bg})

def build_table_html(rows, has_price=False):
    thead_cols = ["Vendor","Category","File Name"]
    if has_price:
        thead_cols += ["Quoted Price"]
    thead_cols += ["Extracted Price","Price Score","Verdict","Link"]

    header = html.Thead(
        html.Tr([html.Th(c) for c in thead_cols]),
        style={"background":PWC["dark"]})

    tbody_rows = []
    for i, r in enumerate(rows):
        bg = "white" if i % 2 == 0 else PWC["light"]
        v_color = r.get("v_color", PWC["grey"])
        cells   = [
            html.Td(vendor_pill(r["vendor"], v_color)),
            html.Td(r.get("category",""),
                style={"color":"#555"}),
            html.Td(
                html.Span(r.get("fname",""),
                    style={"fontFamily":"monospace",
                           "fontSize":"0.79em",
                           "wordBreak":"break-all"})),
        ]
        if has_price:
            qp = r.get("qp_fmt","—")
            cells.append(html.Td(
                html.Span(qp,
                    style={"color":PWC["green"],
                           "fontWeight":"700",
                           "fontFamily":"monospace"}
                    if qp != "—"
                    else html.Span("—",
                        style={"color":"#bbb"}))))
        # Extracted price
        ep = r.get("ep_fmt","—")
        cells.append(html.Td(
            html.Span(ep,
                style={"color":PWC["blue"],
                       "fontWeight":"700",
                       "fontFamily":"monospace"})
            if ep != "—"
            else html.Span("—",style={"color":"#bbb"})))

        # Price score
        ps = r.get("price_score")
        if ps is not None:
            cells.append(html.Td(
                html.Span("{}/100".format(ps),
                    style={"fontWeight":"800",
                           "color":score_color(ps)})))
        else:
            cells.append(html.Td(
                html.Span("—",style={"color":"#bbb"})))

        # Verdict
        v = r.get("verdict","—")
        vcolor = (PWC["green"] if v=="Competitive"
                  else PWC["yellow"] if v=="Average"
                  else PWC["red"] if v=="High"
                  else "#bbb")
        cells.append(html.Td(
            html.Span(v,
                style={"color":vcolor,"fontWeight":"700",
                       "fontSize":"0.82em"})))

        # Link
        url = r.get("url","")
        cells.append(html.Td(
            html.A("Open", href=url, target="_blank",
                style={"color":PWC["orange"],
                       "fontWeight":"600",
                       "textDecoration":"none"})
            if url and url.startswith("http")
            else html.Span("—",style={"color":"#bbb"})))

        tbody_rows.append(html.Tr(cells,
            style={"background":bg}))

    return html.Table(
        [header, html.Tbody(tbody_rows)],
        className="pwc-table")

# ════════════════════════════════════════════════════════════
# LAYOUT COMPONENTS
# ════════════════════════════════════════════════════════════
def build_sidebar(df_master, df_exploded, vcmap):
    cats    = sorted([c for c in df_master["Category"].unique()
                     if str(c).strip() not in ["","nan"]])
    vendors = sorted([v for v in df_master["Vendor"].unique()
                     if str(v).strip() not in ["","nan"]])
    all_svcs = sorted([s for s in df_exploded["Service"].unique()
                       if str(s).strip() not in ["","nan"]])
    groups   = group_services(all_svcs)

    group_items = []
    for gname, gsvcs in groups.items():
        group_items.append(
            dbc.AccordionItem(
                children=[
                    dbc.Row([
                        dbc.Col(
                            dbc.Button(
                                "All",
                                id={"type":"all-btn","group":gname},
                                color="danger",
                                size="sm",
                                style={"width":"100%",
                                       "fontSize":"0.72em"}),
                            width=6),
                        dbc.Col(
                            dbc.Button(
                                "Clear",
                                id={"type":"clr-btn","group":gname},
                                color="secondary",
                                size="sm",
                                style={"width":"100%",
                                       "fontSize":"0.72em"}),
                            width=6),
                    ], className="g-1 mb-2"),
                    dbc.Checklist(
                        options=[{"label":s,"value":s}
                                 for s in gsvcs],
                        value=[],
                        id={"type":"svc-check","group":gname},
                        style={"color":"#ddd",
                               "fontSize":"0.78em"},
                    ),
                ],
                title="{} ({})".format(gname, len(gsvcs)),
                item_id=gname,
            ))

    return html.Div([
        # Logo
        html.Div([
            html.Div("📋", style={"fontSize":"2em"}),
            html.Div("IT Procurement",
                style={"fontSize":"1.0em","fontWeight":"700",
                       "color":"white","margin":"5px 0 2px"}),
            html.Div("Intelligence Dashboard",
                style={"fontSize":"0.68em","color":"#aaa",
                       "letterSpacing":"1px",
                       "textTransform":"uppercase"}),
        ], style={"textAlign":"center","padding":"20px 16px 14px",
                  "borderBottom":"2px solid #D04A02",
                  "marginBottom":"14px"}),

        # Category
        html.Div("📂 Category", className="sb-label"),
        dcc.Dropdown(
            id="sb-cat",
            options=[{"label":"All","value":"All"}] +
                    [{"label":c,"value":c} for c in cats],
            value="All",
            clearable=False,
            style={"margin":"4px 16px 8px",
                   "fontSize":"0.85em"}),

        # Vendor
        html.Div("🏢 Vendor", className="sb-label"),
        dcc.Dropdown(
            id="sb-vendor",
            options=[{"label":"All","value":"All"}] +
                    [{"label":v,"value":v} for v in vendors],
            value="All",
            clearable=False,
            style={"margin":"4px 16px 8px",
                   "fontSize":"0.85em"}),

        html.Hr(style={"borderColor":"#555","margin":"12px 16px"}),

        # Search
        html.Div("🔍 Search Services", className="sb-label"),
        dcc.Input(
            id="sb-search",
            placeholder="e.g. Cisco, Oracle…",
            debounce=True,
            style={"width":"calc(100% - 32px)",
                   "margin":"4px 16px 8px",
                   "padding":"7px 10px",
                   "border":"1px solid #999",
                   "borderRadius":"2px",
                   "background":"white",
                   "fontSize":"0.85em",
                   "fontFamily":"inherit"}),

        # Service groups
        html.Div("🛠 Services by Group", className="sb-label"),
        html.Div(
            dbc.Accordion(
                group_items,
                id="svc-accordion",
                always_open=True,
                flush=True,
                style={"margin":"4px 16px 8px",
                       "--bs-accordion-bg":"rgba(255,255,255,0.05)",
                       "--bs-accordion-color":"#F0F0F0",
                       "--bs-accordion-border-color":"#555"},
            )
        ),

        html.Hr(style={"borderColor":"#555","margin":"12px 16px"}),

        # Stats
        html.Div(id="sb-stats",
            style={"padding":"4px 16px",
                   "fontSize":"0.75em","color":"#888"}),
        html.Div(id="sb-sel-count",
            style={"padding":"4px 16px",
                   "fontSize":"0.78em","fontWeight":"700",
                   "color":PWC["orange"]}),

        html.Hr(style={"borderColor":"#555","margin":"12px 16px"}),

        html.Div(
            dbc.Button(
                "🔄 Reset to Default Catalog",
                id="reset-catalog-btn",
                color="secondary",
                size="sm",
                style={"width":"100%",
                       "fontFamily":"inherit"}),
            style={"padding":"0 16px"}),

    ], className="sidebar")


def build_main_content():
    return html.Div([
        # Header
        html.Div([
            html.Div("IT Procurement Analytics",
                style={"fontSize":"0.68em","fontWeight":"700",
                       "letterSpacing":"2px",
                       "textTransform":"uppercase",
                       "color":PWC["orange"],
                       "marginBottom":"5px"}),
            html.H1("Procurement Intelligence Dashboard",
                style={"margin":"0","fontSize":"1.45em",
                       "fontWeight":"700","color":"white"}),
            html.P(
                "Browse quotations · Compare prices · "
                "Upload & score · AI insights · Catalog upload",
                style={"margin":"6px 0 0","opacity":"0.6",
                       "fontSize":"0.85em","color":"white"}),
        ], style={"background":PWC["dark"],"color":"white",
                  "padding":"20px 28px","borderRadius":"4px",
                  "borderLeft":"6px solid {}".format(PWC["orange"]),
                  "marginBottom":"22px"}),

        # KPIs
        html.Div(id="kpi-row",
            style={"display":"grid",
                   "gridTemplateColumns":"repeat(4,1fr)",
                   "gap":"12px","marginBottom":"20px"}),

        # Tabs
        dbc.Tabs([
            dbc.Tab(label="📊 Analytics",    tab_id="analytics"),
            dbc.Tab(label="📋 Browse",       tab_id="browse"),
            dbc.Tab(label="📤 Upload",       tab_id="upload"),
            dbc.Tab(label="📄 Data Table",   tab_id="table"),
            dbc.Tab(label="🗂 Upload Catalog", tab_id="catalog"),
        ], id="main-tabs", active_tab="analytics",
           style={"marginBottom":"20px"}),

        html.Div(id="tab-content"),

    ], style={"flex":"1","padding":"20px 24px","overflowY":"auto"})


# ════════════════════════════════════════════════════════════
# APP LAYOUT
# ════════════════════════════════════════════════════════════
app.layout = html.Div([
    # Inject CSS
    html.Div(
        dangerouslySetInnerHTML={"__html": CUSTOM_CSS}),

    # Data stores
    dcc.Store(id="store-master",
        data=df_default.to_dict("records")),
    dcc.Store(id="store-exploded",
        data=df_exp_default.to_dict("records")),
    dcc.Store(id="store-selected-svcs", data=[]),
    dcc.Store(id="store-price-cache", data={}),
    dcc.Store(id="store-new-catalog-master", data=None),
    dcc.Store(id="store-new-catalog-exp",    data=None),
    dcc.Store(id="store-quot-result",        data=None),

    # App shell
    html.Div([
        html.Div(id="sidebar-container"),
        build_main_content(),
    ], style={"display":"flex","minHeight":"100vh"}),

], style={"fontFamily":"Source Sans Pro,Helvetica Neue,Arial,sans-serif",
          "background":PWC["light"]})


# ════════════════════════════════════════════════════════════
# CALLBACKS
# ════════════════════════════════════════════════════════════

# ── Render sidebar ──────────────────────────────────────────
@app.callback(
    Output("sidebar-container","children"),
    Input("store-master","data"),
    Input("store-exploded","data"),
)
def render_sidebar(master_data, exp_data):
    df_m = pd.DataFrame(master_data or [])
    df_e = pd.DataFrame(exp_data   or [])
    if df_m.empty:
        return html.Div("No data loaded.",
            className="sidebar",
            style={"padding":"20px","color":"white"})
    vcmap = {v:get_color(i)
             for i,v in enumerate(
                 sorted(df_m["Vendor"].unique()))}
    return build_sidebar(df_m, df_e, vcmap)


# ── Collect selected services ───────────────────────────────
@app.callback(
    Output("store-selected-svcs","data"),
    Output("sb-sel-count","children"),
    Input({"type":"svc-check","group":dash.ALL},"value"),
)
def collect_selected_svcs(all_values):
    selected = []
    for vals in (all_values or []):
        if vals: selected.extend(vals)
    selected = list(set(selected))
    count_txt = (
        "✅ {} service(s) selected".format(len(selected))
        if selected else "")
    return selected, count_txt


# ── Update stats ────────────────────────────────────────────
@app.callback(
    Output("sb-stats","children"),
    Input("store-master","data"),
    Input("store-exploded","data"),
)
def update_stats(master_data, exp_data):
    df_m = pd.DataFrame(master_data or [])
    df_e = pd.DataFrame(exp_data   or [])
    return [
        "📄 {} quotes".format(
            df_m["File Name"].nunique() if not df_m.empty else 0),
        html.Br(),
        "🛠 {} services".format(
            df_e["Service"].nunique() if not df_e.empty else 0),
        html.Br(),
        "🏢 {} vendors".format(
            df_m["Vendor"].nunique() if not df_m.empty else 0),
    ]


# ── KPI row ─────────────────────────────────────────────────
@app.callback(
    Output("kpi-row","children"),
    Input("store-master","data"),
    Input("store-exploded","data"),
    Input("sb-cat","data"),
    Input("sb-vendor","data"),
)
def update_kpis(master_data, exp_data, cat, vendor):
    df_m = pd.DataFrame(master_data or [])
    df_e = pd.DataFrame(exp_data   or [])
    if df_m.empty:
        return []
    if cat and cat != "All":
        df_e = df_e[df_e["Category"]==cat]
        df_m = df_m[df_m["Category"]==cat]
    if vendor and vendor != "All":
        df_e = df_e[df_e["Vendor"]==vendor]
        df_m = df_m[df_m["Vendor"]==vendor]
    return [
        kpi_card(df_m["File Name"].nunique(),
                 "Total Quotes",    PWC["orange"]),
        kpi_card(df_e["Service"].nunique() if not df_e.empty else 0,
                 "Unique Services", PWC["blue"]),
        kpi_card(df_m["Vendor"].nunique(),
                 "Vendors",         PWC["teal"]),
        kpi_card(df_m["Category"].nunique(),
                 "Categories",      PWC["dark"]),
    ]


# ── Tab content ─────────────────────────────────────────────
@app.callback(
    Output("tab-content","children"),
    Input("main-tabs","active_tab"),
    Input("store-master","data"),
    Input("store-exploded","data"),
    Input("store-selected-svcs","data"),
    Input("store-price-cache","data"),
    Input("store-quot-result","data"),
    Input("sb-cat","data"),
    Input("sb-vendor","data"),
)
def render_tab(
    active_tab, master_data, exp_data,
    selected_svcs, price_cache, quot_result,
    cat, vendor
):
    df_m = pd.DataFrame(master_data or [])
    df_e = pd.DataFrame(exp_data   or [])

    if df_m.empty and active_tab != "catalog":
        return dbc.Alert(
            "No catalog loaded. Go to 🗂 Upload Catalog tab.",
            color="info")

    # Apply sidebar filters
    df_ef = df_e.copy()
    df_mf = df_m.copy()
    if cat and cat != "All":
        df_ef = df_ef[df_ef["Category"]==cat]
        df_mf = df_mf[df_mf["Category"]==cat]
    if vendor and vendor != "All":
        df_ef = df_ef[df_ef["Vendor"]==vendor]
        df_mf = df_mf[df_mf["Vendor"]==vendor]

    vcmap = {v:get_color(i)
             for i,v in enumerate(
                 sorted(df_m["Vendor"].unique()))}

    # ── ANALYTICS ──────────────────────────────────────────
    if active_tab == "analytics":
        return html.Div([
            dbc.Row([
                dbc.Col([
                    section_title_div(
                        "SERVICE OVERLAP ANALYSIS",
                        "Orange = service quoted by multiple vendors."),
                    dcc.Graph(
                        figure=build_overlap_chart(df_ef),
                        config={"displayModeBar":False}),
                ], width=6),
                dbc.Col([
                    section_title_div(
                        "VENDOR SERVICE COVERAGE",
                        "Higher = broader vendor capability."),
                    dcc.Graph(
                        figure=build_coverage_chart(df_ef, vcmap),
                        config={"displayModeBar":False}),
                ], width=6),
            ], className="mb-4"),
            html.Div([
                section_title_div(
                    "PROCUREMENT CATEGORY DISTRIBUTION",
                    "Share of quote files across categories."),
                dcc.Graph(
                    figure=build_category_chart(df_ef),
                    config={"displayModeBar":False}),
            ], style={"background":"white","padding":"16px",
                      "borderRadius":"4px",
                      "border":"1px solid #e0e0e0",
                      "marginBottom":"20px"}),
            ai_box_div(
                ai_service_summary(df_m, df_e)
                if not df_m.empty else "No data."),
        ])

    # ── BROWSE ─────────────────────────────────────────────
    elif active_tab == "browse":
        if not selected_svcs:
            return dbc.Alert(
                "👈 Select services from the sidebar to browse.",
                color="info")

        d_sel = df_ef[df_ef["Service"].isin(selected_svcs)].copy()
        if d_sel.empty:
            return dbc.Alert(
                "No results found under current filters.",
                color="warning")

        vsmap = defaultdict(set)
        for _, r in d_sel.iterrows():
            vsmap[r["Vendor"]].add(r["Service"])
        vendors_all  = sorted([v for v,s in vsmap.items()
                                if set(selected_svcs).issubset(s)])
        vendors_some = sorted([v for v,s in vsmap.items()
                                if not set(selected_svcs).issubset(s)])

        content = []

        # Summary banners
        if len(selected_svcs) > 1:
            if vendors_all:
                content.append(dbc.Alert(
                    "✅ {} vendor(s) cover ALL {} services: {}".format(
                        len(vendors_all), len(selected_svcs),
                        " · ".join(vendors_all)),
                    color="success"))
            else:
                content.append(dbc.Alert(
                    "⚠️ No single vendor covers all {} services.".format(
                        len(selected_svcs)),
                    color="warning"))

        content.append(section_title_div(
            "QUOTATION FILES — PER SERVICE",
            "Price Score: 100=cheapest, 0=most expensive."))

        has_price = "Quoted Price" in d_sel.columns

        for svc in selected_svcs:
            d_svc = (d_sel[d_sel["Service"]==svc]
                     .drop_duplicates(subset=["Vendor","File Name"])
                     .sort_values("Vendor"))
            vc    = d_svc["Vendor"].nunique()
            s_tag = "SHARED" if vc > 1 else "SINGLE VENDOR"

            # Collect prices
            all_prices = []
            for _, r in d_svc.iterrows():
                qp = _parse_num(str(r.get("Quoted Price","")).strip())
                if qp > 0: all_prices.append(qp)
                ck = "px_{}".format(str(r.get("File Name","")).strip())
                ca = (price_cache or {}).get(ck)
                if ca and ca.get("price_num",0) > 0:
                    all_prices.append(ca["price_num"])

            # Build rows
            table_rows = []
            vendor_pmap = {}
            for _, r in d_svc.iterrows():
                fname  = str(r.get("File Name","")).strip()
                url    = str(r.get("Hyperlink","")).strip()
                if not url or url=="nan":
                    url = str(r.get("File Link","")).strip()
                if url == "nan": url = ""
                qp_str = str(r.get("Quoted Price","")).strip()
                qp_num = _parse_num(qp_str)
                ck     = "px_{}".format(fname)
                cached = (price_cache or {}).get(ck)
                ep_num = cached["price_num"] if cached and cached.get("price_num",0)>0 else 0
                ep_fmt = _fmt(cached["price"]) if ep_num>0 else "—"
                ref    = ep_num if ep_num>0 else qp_num
                vendor_pmap[r["Vendor"]] = ref
                others = [p for p in all_prices if p != ref]
                ps     = None
                verdict= "—"
                if ref > 0 and others:
                    ps, _, _, _, _ = price_score(ref, others)
                    if ps is not None:
                        verdict = ("Competitive" if ps>=70
                                   else "Average" if ps>=40
                                   else "High")
                table_rows.append({
                    "vendor"     : r["Vendor"],
                    "v_color"    : vcmap.get(r["Vendor"],PWC["grey"]),
                    "category"   : r.get("Category",""),
                    "fname"      : fname,
                    "qp_fmt"     : _fmt(qp_str) if qp_num>0 else "—",
                    "ep_fmt"     : ep_fmt,
                    "price_score": ps,
                    "verdict"    : verdict,
                    "url"        : url,
                })

            # Price chart
            price_data_chart = []
            for _, r in d_svc.iterrows():
                fname  = str(r.get("File Name","")).strip()
                ck     = "px_{}".format(fname)
                cached = (price_cache or {}).get(ck)
                qp     = _parse_num(str(r.get("Quoted Price","")).strip())
                ep     = cached["price_num"] if cached else 0.0
                pval   = ep if ep>0 else qp
                if pval > 0:
                    price_data_chart.append({
                        "Label" : "{}/{}".format(r["Vendor"],fname[:15]),
                        "Price" : pval,
                        "Type"  : "Historical",
                        "Color" : vcmap.get(r["Vendor"],PWC["grey"]),
                    })
            avg_p = (sum(x["Price"] for x in price_data_chart) /
                     len(price_data_chart)) if price_data_chart else 0

            ai_txt = ai_price_insight(
                0, all_prices,
                {v:p for v,p in vendor_pmap.items() if p>0})

            content.append(
                dbc.Accordion([
                    dbc.AccordionItem(
                        children=[
                            html.Div([
                                html.B("Vendors: ",
                                    style={"fontSize":"0.87em"}),
                                *[vendor_pill(v,vcmap.get(v,PWC["grey"]))
                                  for v in sorted(d_svc["Vendor"].unique())]
                            ], style={"marginBottom":"12px"}),
                            html.Div(
                                build_table_html(
                                    table_rows, has_price),
                                style={"overflowX":"auto"}),
                            html.Br(),
                            ai_box_div(ai_txt) if all_prices else html.Div(),
                            html.Br(),
                            dbc.Button(
                                "Extract Prices — {}".format(svc[:40]),
                                id={"type":"extract-btn","svc":svc},
                                color="danger",
                                style={"fontFamily":"inherit",
                                       "fontWeight":"700"}),
                            html.Br(), html.Br(),
                            dcc.Graph(
                                figure=build_price_chart(
                                    price_data_chart, avg_p),
                                config={"displayModeBar":False})
                            if len(price_data_chart)>=2
                            else html.Div(),
                        ],
                        title="{}  —  {} vendor(s) · {} file(s) · {}".format(
                            svc, vc, len(d_svc), s_tag),
                    )
                ], always_open=True, flush=True,
                   style={"marginBottom":"10px"}))

        return html.Div(content)

    # ── UPLOAD & SCORE ─────────────────────────────────────
    elif active_tab == "upload":
        content = [
            html.Div([
                html.Div("New Quotation Analysis",
                    style={"fontSize":"0.68em","fontWeight":"700",
                           "letterSpacing":"2px",
                           "textTransform":"uppercase",
                           "color":PWC["orange"],
                           "marginBottom":"4px"}),
                html.Div(
                    "Upload a quotation → score vs history",
                    style={"fontSize":"1.0em","fontWeight":"700",
                           "color":"white"}),
                html.P(
                    "PDF, XLSX, DOCX · Price Score 0-100",
                    style={"opacity":"0.6","fontSize":"0.83em",
                           "color":"white","marginTop":"4px"}),
            ], style={"background":PWC["dark"],"padding":"14px 20px",
                      "borderRadius":"4px",
                      "borderLeft":"6px solid {}".format(PWC["orange"]),
                      "marginBottom":"16px"}),

            # Upload
            html.Div([
                html.Div("Step 1 — Upload File",
                    className="step-num"),
                dcc.Upload(
                    id="upload-quot",
                    children=html.Div([
                        html.Div("📄",
                            style={"fontSize":"2em",
                                   "marginBottom":"8px"}),
                        html.Div(
                            "Drag & drop or click to upload",
                            style={"fontWeight":"700",
                                   "color":PWC["orange"]}),
                        html.Div(
                            "PDF, Excel, Word",
                            style={"fontSize":"0.80em",
                                   "color":"#999",
                                   "marginTop":"4px"}),
                    ]),
                    className="upload-zone",
                    multiple=False),
            ], className="step-box"),
        ]

        # Show extraction result
        if quot_result:
            np = quot_result.get("price_num", 0)
            content.append(html.Div([
                html.Div("Step 2 — Extracted Price",
                    className="step-num"),
                html.Div([
                    html.Div("Extracted Price",
                        style={"fontSize":"0.68em",
                               "fontWeight":"700",
                               "color":PWC["green"],
                               "textTransform":"uppercase",
                               "letterSpacing":"1px"}),
                    html.Div(
                        _fmt(np) if np>0 else "Not found",
                        style={"fontSize":"2.0em",
                               "fontWeight":"800",
                               "color":PWC["green"] if np>0
                               else PWC["grey"]}),
                ], className="score-card green"
                   if np>0 else "score-card yellow"),
                dcc.Input(
                    id="manual-price",
                    type="number",
                    placeholder="Enter price manually if needed",
                    style={"border":"1px solid #ccc",
                           "padding":"6px 10px",
                           "borderRadius":"2px",
                           "width":"280px",
                           "fontFamily":"inherit",
                           "marginTop":"8px"}),
            ], className="step-box"))

            # Service selection
            all_svcs_up = sorted([s for s in df_e["Service"].unique()
                                   if str(s).strip() not in ["","nan"]])
            svc_groups_up = group_services(all_svcs_up)
            svc_checks_up = []
            for gname, gsvcs in svc_groups_up.items():
                svc_checks_up.append(
                    dbc.AccordionItem(
                        dbc.Checklist(
                            options=[{"label":s,"value":s}
                                     for s in gsvcs],
                            value=[],
                            id={"type":"up-svc-check",
                                "group":gname},
                            style={"fontSize":"0.80em"}),
                        title="{} ({})".format(gname,len(gsvcs)),
                        item_id="up_{}".format(gname)))

            content.append(html.Div([
                html.Div("Step 3 — Select Services",
                    className="step-num"),
                dbc.Accordion(
                    svc_checks_up,
                    always_open=True, flush=True),
                dcc.Dropdown(
                    id="cat-filter-up",
                    options=[{"label":"All","value":"All"}] +
                            [{"label":c,"value":c}
                             for c in sorted(
                                 df_m["Category"].unique())],
                    value="All",
                    clearable=False,
                    placeholder="Filter by category",
                    style={"marginTop":"10px",
                           "fontSize":"0.85em"}),
            ], className="step-box"))

            content.append(
                dbc.Button(
                    "Run Comparison →",
                    id="run-compare",
                    color="danger",
                    style={"fontFamily":"inherit",
                           "fontWeight":"700",
                           "marginBottom":"12px"}))

        content.append(html.Div(id="compare-results"))
        return html.Div(content)

    # ── DATA TABLE ─────────────────────────────────────────
    elif active_tab == "table":
        cols = [c for c in df_mf.columns
                if c not in ["Services List","Hyperlink"]]
        return html.Div([
            section_title_div("ALL QUOTATION DATA"),
            html.Div(style={"overflowX":"auto"},
                children=[
                    dash_table.DataTable(
                        data=df_mf[cols].to_dict("records"),
                        columns=[{"name":c,"id":c} for c in cols],
                        style_table={"overflowX":"auto"},
                        style_cell={
                            "fontFamily":"Source Sans Pro,sans-serif",
                            "fontSize":"0.82em",
                            "padding":"8px 10px",
                            "textAlign":"left",
                            "wordBreak":"break-word",
                            "minWidth":"80px",
                            "maxWidth":"300px",
                        },
                        style_header={
                            "background":PWC["dark"],
                            "color":"white",
                            "fontWeight":"700",
                            "fontSize":"0.78em",
                            "textTransform":"uppercase",
                            "letterSpacing":"0.4px",
                            "border":"none",
                        },
                        style_data_conditional=[
                            {"if":{"row_index":"odd"},
                             "backgroundColor":PWC["light"]},
                        ],
                        filter_action="native",
                        sort_action="native",
                        page_size=20,
                        page_action="native",
                    )
                ])
        ])

    # ── CATALOG UPLOAD ─────────────────────────────────────
    elif active_tab == "catalog":
        return html.Div([
            html.Div([
                html.Div("Catalog Management",
                    style={"fontSize":"0.68em","fontWeight":"700",
                           "letterSpacing":"2px",
                           "textTransform":"uppercase",
                           "color":PWC["orange"],
                           "marginBottom":"5px"}),
                html.H1("Upload Master Catalog",
                    style={"fontSize":"1.2em","fontWeight":"700",
                           "color":"white","margin":"0"}),
                html.P(
                    "Upload any Excel or CSV catalog — "
                    "AI auto-detects columns and builds "
                    "the dashboard automatically.",
                    style={"opacity":"0.6","fontSize":"0.85em",
                           "color":"white","marginTop":"4px"}),
            ], style={"background":PWC["dark"],
                      "padding":"18px 24px",
                      "borderRadius":"4px",
                      "borderLeft":"6px solid {}".format(
                          PWC["orange"]),
                      "marginBottom":"20px"}),

            html.Div(id="catalog-current-status"),

            section_title_div("UPLOAD YOUR CATALOG FILE"),
            dcc.Upload(
                id="upload-catalog",
                children=html.Div([
                    html.Div("📊",
                        style={"fontSize":"2.5em",
                               "marginBottom":"10px"}),
                    html.Div(
                        "Drag & drop or click to upload",
                        style={"fontWeight":"700",
                               "color":PWC["orange"],
                               "fontSize":"1.0em"}),
                    html.Div(
                        "Excel (.xlsx / .xls) or CSV",
                        style={"fontSize":"0.80em",
                               "color":"#999",
                               "marginTop":"4px"}),
                ]),
                className="upload-zone",
                multiple=False,
                style={"marginTop":"10px",
                       "marginBottom":"16px"}),

            html.Div(id="catalog-process-area"),
            html.Div(id="catalog-apply-area"),
        ])

    return html.Div("Select a tab.")


# ── Handle quotation file upload ────────────────────────────
@app.callback(
    Output("store-quot-result","data"),
    Input("upload-quot","contents"),
    State("upload-quot","filename"),
    prevent_initial_call=True,
)
def handle_quot_upload(contents, filename):
    if not contents or not filename:
        return None
    content_type, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)
    ext     = filename.rsplit(".",1)[-1].lower()
    result  = extract_price_from_bytes(decoded, ext)
    result["filename"] = filename
    return result


# ── Run comparison ──────────────────────────────────────────
@app.callback(
    Output("compare-results","children"),
    Input("run-compare","n_clicks"),
    State("store-quot-result","data"),
    State("store-master","data"),
    State("store-exploded","data"),
    State({"type":"up-svc-check","group":dash.ALL},"value"),
    State("cat-filter-up","value"),
    State("manual-price","value"),
    State("store-price-cache","data"),
    prevent_initial_call=True,
)
def run_comparison(
    n_clicks, quot_result, master_data, exp_data,
    up_svc_values, cat_filter, manual_price, price_cache
):
    if not n_clicks: return html.Div()
    df_m = pd.DataFrame(master_data or [])
    df_e = pd.DataFrame(exp_data   or [])
    if df_m.empty: return dbc.Alert("No catalog data.",color="warning")

    new_price = (quot_result or {}).get("price_num", 0)
    new_text  = (quot_result or {}).get("text", "")
    fname_up  = (quot_result or {}).get("filename", "")
    if manual_price and float(manual_price or 0) > 0:
        new_price = float(manual_price)

    new_services = []
    for vals in (up_svc_values or []):
        if vals: new_services.extend(vals)

    if not new_services and new_price <= 0:
        return dbc.Alert("Select services and/or provide a price.",
                         color="info")

    candidates = (df_e[df_e["Service"].isin(new_services)]
                  if new_services else df_e.copy())
    if cat_filter and cat_filter != "All":
        candidates = candidates[
            candidates["Category"]==cat_filter]

    cand_files = (candidates
                  .drop_duplicates(subset=["File Name","Vendor"])
                  [["File Name","Vendor","Category",
                    "Hyperlink","Quoted Price"]]
                  .copy())

    if cand_files.empty:
        return dbc.Alert("No historical quotes found.",
                         color="warning")

    hist_prices = []
    vendor_pmap = {}
    for _, r in cand_files.iterrows():
        qp = _parse_num(str(r.get("Quoted Price","")).strip())
        if qp > 0:
            hist_prices.append(qp)
            vendor_pmap[r["Vendor"]] = qp
        ck = "px_{}".format(str(r.get("File Name","")).strip())
        ca = (price_cache or {}).get(ck)
        if ca and ca.get("price_num",0) > 0:
            hist_prices.append(ca["price_num"])
            vendor_pmap[r["Vendor"]] = ca["price_num"]

    content = []

    if new_price > 0 and hist_prices:
        ps, ps_label, avg_h, mn_h, mx_h = price_score(
            new_price, hist_prices)
        sc = score_color(ps)
        css = score_css(ps)

        content.append(html.Div([
            html.Div("Step 4 — Comparison Results",
                className="step-num"),
            dbc.Row([
                dbc.Col(score_card_div(
                    "Price Score",
                    "{}/100".format(ps) if ps else "N/A",
                    "vs {} quotes".format(len(hist_prices)),
                    css, sc), width=3),
                dbc.Col(score_card_div(
                    "Your Price", _fmt(new_price),
                    fname_up[:30],
                    "yellow", PWC["orange"]), width=3),
                dbc.Col(score_card_div(
                    "Historical Avg", _fmt(avg_h),
                    "min {} · max {}".format(
                        _fmt(mn_h), _fmt(mx_h)),
                    "yellow", PWC["blue"]), width=3),
                dbc.Col(score_card_div(
                    "Verdict", ps_label,
                    "", css, sc), width=3),
            ], className="mb-3"),
        ], className="step-box"))

        content.append(ai_box_div(
            ai_price_insight(new_price, hist_prices, vendor_pmap)))

        # Price chart
        chart_data = []
        for _, r in cand_files.iterrows():
            fn     = str(r.get("File Name","")).strip()
            ck     = "px_{}".format(fn)
            cached = (price_cache or {}).get(ck)
            qp     = _parse_num(str(r.get("Quoted Price","")).strip())
            ep     = cached["price_num"] if cached else 0.0
            pval   = ep if ep>0 else qp
            if pval > 0:
                chart_data.append({
                    "Label" : "{}/{}".format(r["Vendor"],fn[:12]),
                    "Price" : pval,
                    "Type"  : "Historical",
                    "Color" : get_color(0),
                })
        chart_data.append({
            "Label" : "NEW:{}".format(fname_up[:15]),
            "Price" : new_price,
            "Type"  : "New",
            "Color" : PWC["orange"],
        })
        content.append(html.Div([
            section_title_div("PRICE POSITIONING"),
            dcc.Graph(
                figure=build_price_chart(chart_data, avg_h),
                config={"displayModeBar":False}),
        ], style={"background":"white","padding":"16px",
                  "borderRadius":"4px",
                  "border":"1px solid #e0e0e0",
                  "marginTop":"16px"}))

        # Similarity table
        sim_rows = []
        for _, r in cand_files.iterrows():
            fn      = str(r.get("File Name","")).strip()
            hist_s  = list(df_e[df_e["File Name"]==fn
                               ]["Service"].unique())
            svc_sim = service_similarity(new_services, hist_s)
            sim_rows.append({
                "Vendor"   : r["Vendor"],
                "Category" : r.get("Category",""),
                "File"     : fn,
                "Svc %"    : svc_sim,
                "Score"    : svc_sim,
                "URL"      : str(r.get("Hyperlink","")).strip(),
            })

        if sim_rows:
            sim_df = pd.DataFrame(sim_rows).sort_values(
                "Score", ascending=False)
            sim_table_rows = []
            for i, r in sim_df.iterrows():
                bg  = "white" if i%2==0 else PWC["light"]
                sc2 = score_color(r["Score"])
                url = r["URL"]
                sim_table_rows.append(html.Tr([
                    html.Td(vendor_pill(r["Vendor"],
                        get_color(list(sim_df["Vendor"]
                            .unique()).index(r["Vendor"])))),
                    html.Td(r["Category"],
                        style={"color":"#555"}),
                    html.Td(html.Span(r["File"],
                        style={"fontFamily":"monospace",
                               "fontSize":"0.79em",
                               "wordBreak":"break-all"})),
                    html.Td(html.Span(
                        "{}%".format(r["Svc %"]),
                        style={"color":PWC["blue"],
                               "fontWeight":"700"}),
                        style={"textAlign":"center"}),
                    html.Td(html.Span(
                        "{}/100".format(r["Score"]),
                        style={"color":sc2,
                               "fontWeight":"800",
                               "fontSize":"1.0em"}),
                        style={"textAlign":"center"}),
                    html.Td(html.A(
                        "Open", href=url, target="_blank",
                        style={"color":PWC["orange"],
                               "fontWeight":"600",
                               "textDecoration":"none"})
                        if url and url.startswith("http")
                        else "—"),
                ], style={"background":bg}))

            content.append(html.Div([
                html.Br(),
                section_title_div(
                    "DOCUMENT SIMILARITY",
                    "Service overlap vs historical quotations."),
                html.Div(
                    html.Table([
                        html.Thead(html.Tr([
                            html.Th("Vendor"),
                            html.Th("Category"),
                            html.Th("File"),
                            html.Th("Services %"),
                            html.Th("Score"),
                            html.Th("Link"),
                        ])),
                        html.Tbody(sim_table_rows),
                    ], className="pwc-table"),
                    style={"overflowX":"auto"}),
            ]))

    return html.Div(content)


# ── Handle catalog upload ───────────────────────────────────
@app.callback(
    Output("catalog-process-area","children"),
    Output("catalog-apply-area","children"),
    Output("store-new-catalog-master","data"),
    Output("store-new-catalog-exp","data"),
    Input("upload-catalog","contents"),
    State("upload-catalog","filename"),
    prevent_initial_call=True,
)
def handle_catalog_upload(contents, filename):
    if not contents or not filename:
        return html.Div(), html.Div(), None, None

    content_type, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)

    df_new, df_exp_new, err = process_catalog(decoded, filename)

    if err or df_new is None:
        return (
            dbc.Alert("❌ Error: {}".format(err or "Unknown"),
                      color="danger"),
            html.Div(), None, None)

    insights = ai_analyze_catalog(df_new, df_exp_new)
    vcmap_new = {v:get_color(i)
                 for i,v in enumerate(
                     sorted(df_new["Vendor"].unique()))}

    # Build insight cards
    insight_items = [
        ("Overview",    insights.get("overview",""),    PWC["blue"]),
        ("Top Vendor",  insights.get("top_vendor",""),  PWC["teal"]),
        ("Competitive", insights.get("competitive",""), PWC["orange"]),
        ("Categories",  insights.get("category",""),    PWC["yellow"]),
    ]
    if insights.get("pricing"):
        insight_items.append(
            ("Pricing", insights["pricing"], PWC["green"]))

    insight_cards = []
    for i in range(0, len(insight_items), 2):
        row_items = insight_items[i:i+2]
        insight_cards.append(
            dbc.Row([
                dbc.Col(html.Div([
                    html.Div(title,
                        style={"fontSize":"0.68em",
                               "fontWeight":"700",
                               "letterSpacing":"1px",
                               "textTransform":"uppercase",
                               "color":color,
                               "marginBottom":"6px"}),
                    dcc.Markdown(text,
                        style={"fontSize":"0.85em",
                               "color":PWC["dark"]}),
                ], style={"background":"white",
                          "borderRadius":"4px",
                          "padding":"14px 16px",
                          "border":"1px solid #e0e0e0",
                          "height":"100%"}), width=6)
                for title, text, color in row_items
            ], className="mb-3"))

    recs = insights.get("recommendations",[])
    rec_divs = [html.Div("▸ {}".format(r),
        style={"background":"#FFF3F0",
               "borderLeft":"3px solid {}".format(PWC["orange"]),
               "padding":"8px 12px","borderRadius":"2px",
               "marginBottom":"6px","fontSize":"0.84em"})
        for r in recs]

    # Preview charts
    spv_new = (df_exp_new.groupby("Vendor")["Service"]
               .nunique().sort_values(ascending=False)
               .reset_index())
    spv_new.columns = ["Vendor","Services"]
    vc_map_new = {v:get_color(i)
                  for i,v in enumerate(spv_new["Vendor"].tolist())}

    fig_spv = go.Figure(go.Bar(
        x=spv_new["Vendor"],
        y=spv_new["Services"],
        marker_color=[vc_map_new.get(v,PWC["grey"])
                      for v in spv_new["Vendor"]],
        marker_line_width=0,
        text=spv_new["Services"],
        textposition="outside",
    ))
    fig_spv.update_layout(
        title="Services per Vendor",
        height=320, plot_bgcolor=CBG,
        paper_bgcolor=CBG,
        margin=dict(l=5,r=10,t=40,b=10),
        font=CFONT,
        yaxis=dict(showgrid=True,
                   gridcolor="#E0E0E0",
                   zeroline=False),
        xaxis=dict(tickangle=-30),
        bargap=0.35)

    cat_new = (df_new.drop_duplicates(
                   subset=["Category","File Name"])
               .groupby("Category").size().reset_index())
    cat_new.columns = ["Category","Count"]
    fig_cat = px.pie(
        cat_new, names="Category", values="Count",
        hole=0.45, color_discrete_sequence=PALETTE)
    fig_cat.update_traces(
        textposition="outside",
        textinfo="label+percent",
        textfont_size=10)
    fig_cat.update_layout(
        title="Category Distribution",
        height=320,
        margin=dict(l=10,r=10,t=40,b=10),
        paper_bgcolor=CBG, font=CFONT)

    process_area = html.Div([

        # Step 1 — received
        html.Div([
            html.Div("Step 1 — File Received",
                className="step-num"),
            html.Span("✅ "),
            html.B(filename),
            " — {} KB uploaded successfully.".format(
                round(len(decoded)/1024,1)),
        ], className="step-box"),

        # Step 2 — columns
        html.Div([
            html.Div("Step 2 — AI Column Detection",
                className="step-num"),
            "Mapped columns. Found ",
            html.B("{} rows".format(len(df_new))),
            " across ",
            html.B("{} vendors".format(
                df_new["Vendor"].nunique())),
            " and ",
            html.B("{} categories.".format(
                df_new["Category"].nunique())),
            html.Br(), html.Br(),
            dbc.Row([
                dbc.Col(html.Div([
                    html.Div(col,
                        style={"background":PWC["dark"],
                               "color":PWC["orange"],
                               "padding":"8px 10px",
                               "borderRadius":"2px",
                               "fontSize":"0.80em",
                               "fontWeight":"700",
                               "textAlign":"center",
                               "marginBottom":"4px"}),
                    html.Div(
                        "{} unique".format(
                            df_new[col].nunique()),
                        style={"textAlign":"center",
                               "fontSize":"0.75em",
                               "color":"#555"}),
                ]), width=3)
                for col in ["Category","Vendor",
                             "File Name","Comments"]
                if col in df_new.columns
            ]),
        ], className="step-box"),

        # Step 3 — AI analysis
        html.Div([
            html.Div("Step 3 — AI Analysis",
                className="step-num"),
            html.Div(
                "Generated insights from catalog data:",
                style={"marginBottom":"12px",
                       "fontSize":"0.85em"}),
            *insight_cards,
            html.Div("AI RECOMMENDATIONS",
                className="section-title",
                style={"marginTop":"12px",
                       "marginBottom":"8px"}),
            *rec_divs if rec_divs
            else [html.Div(
                "No specific recommendations.",
                style={"color":"#888",
                       "fontSize":"0.84em"})],
        ], className="step-box"),

        # Step 4 — Preview charts
        html.Div([
            html.Div("Step 4 — Catalog Preview",
                className="step-num"),
            dbc.Row([
                dbc.Col(
                    dcc.Graph(
                        figure=fig_spv,
                        config={"displayModeBar":False}),
                    width=6),
                dbc.Col(
                    dcc.Graph(
                        figure=fig_cat,
                        config={"displayModeBar":False}),
                    width=6),
            ]),
        ], className="step-box"),

        # Step 5 — Data preview table
        html.Div([
            html.Div("Step 5 — Data Preview "
                     "(First 20 rows)",
                className="step-num"),
            html.Div(
                dash_table.DataTable(
                    data=(df_new.drop(
                        columns=["Services List",
                                 "Hyperlink"],
                        errors="ignore")
                        .head(20)
                        .to_dict("records")),
                    columns=[
                        {"name":c,"id":c}
                        for c in df_new.columns
                        if c not in [
                            "Services List","Hyperlink"]
                    ],
                    style_table={"overflowX":"auto"},
                    style_cell={
                        "fontFamily":
                            "Source Sans Pro,sans-serif",
                        "fontSize":"0.80em",
                        "padding":"7px 10px",
                        "textAlign":"left",
                        "wordBreak":"break-word",
                        "minWidth":"80px",
                        "maxWidth":"250px",
                    },
                    style_header={
                        "background":PWC["dark"],
                        "color":"white",
                        "fontWeight":"700",
                        "fontSize":"0.76em",
                        "textTransform":"uppercase",
                        "letterSpacing":"0.4px",
                        "border":"none",
                    },
                    style_data_conditional=[
                        {"if":{"row_index":"odd"},
                         "backgroundColor":PWC["light"]},
                    ],
                    page_size=20,
                ),
                style={"overflowX":"auto"}),
        ], className="step-box"),

    ])

    apply_area = html.Div([
        html.Br(),
        dbc.Button(
            "✅ Apply This Catalog to Dashboard",
            id="apply-catalog-btn",
            color="danger",
            size="lg",
            style={"fontFamily":"inherit",
                   "fontWeight":"700",
                   "letterSpacing":"0.3px"}),
        html.Div(id="apply-catalog-status",
            style={"marginTop":"10px"}),
    ])

    return (
        process_area,
        apply_area,
        df_new.to_dict("records"),
        df_exp_new.to_dict("records"),
    )


# ── Apply catalog to dashboard ──────────────────────────────
@app.callback(
    Output("store-master","data"),
    Output("store-exploded","data"),
    Output("apply-catalog-status","children"),
    Input("apply-catalog-btn","n_clicks"),
    State("store-new-catalog-master","data"),
    State("store-new-catalog-exp","data"),
    prevent_initial_call=True,
)
def apply_catalog(n_clicks, new_master, new_exp):
    if not n_clicks or not new_master:
        return dash.no_update, dash.no_update, html.Div()
    df_new     = pd.DataFrame(new_master)
    df_exp_new = pd.DataFrame(new_exp or [])
    status = dbc.Alert(
        "✅ Catalog applied! Dashboard updated with "
        "{} vendors, {} services, {} categories.".format(
            df_new["Vendor"].nunique(),
            df_exp_new["Service"].nunique()
            if not df_exp_new.empty else 0,
            df_new["Category"].nunique()),
        color="success",
        style={"fontSize":"0.88em"})
    return (
        df_new.to_dict("records"),
        df_exp_new.to_dict("records"),
        status,
    )


# ── Reset to default catalog ────────────────────────────────
@app.callback(
    Output("store-master","data",
           allow_duplicate=True),
    Output("store-exploded","data",
           allow_duplicate=True),
    Input("reset-catalog-btn","n_clicks"),
    prevent_initial_call=True,
)
def reset_catalog(n_clicks):
    if not n_clicks:
        return dash.no_update, dash.no_update
    df_m, df_e = load_default()
    if df_m is None:
        return (
            pd.DataFrame(columns=[
                "Category","Vendor","File Name",
                "Comments","Hyperlink",
                "Services List"]).to_dict("records"),
            pd.DataFrame(columns=[
                "Category","Vendor","File Name",
                "Comments","Hyperlink",
                "Service"]).to_dict("records"),
        )
    return (
        df_m.to_dict("records"),
        df_e.to_dict("records"),
    )


# ── Catalog current status banner ───────────────────────────
@app.callback(
    Output("catalog-current-status","children"),
    Input("store-master","data"),
    Input("store-exploded","data"),
)
def update_catalog_status(master_data, exp_data):
    df_m = pd.DataFrame(master_data or [])
    df_e = pd.DataFrame(exp_data   or [])
    if df_m.empty:
        return dbc.Alert(
            "No catalog loaded.",
            color="warning",
            style={"fontSize":"0.85em"})
    return dbc.Alert(
        "✅ Active catalog: {} rows · {} vendors · "
        "{} services · {} categories".format(
            len(df_m),
            df_m["Vendor"].nunique(),
            df_e["Service"].nunique()
            if not df_e.empty else 0,
            df_m["Category"].nunique()),
        color="success",
        style={"fontSize":"0.85em",
               "marginBottom":"14px"})


# ── Vendor dropdown updates on category change ──────────────
@app.callback(
    Output("sb-vendor","options"),
    Input("sb-cat","value"),
    State("store-master","data"),
)
def update_vendor_dropdown(cat, master_data):
    df_m = pd.DataFrame(master_data or [])
    if df_m.empty:
        return [{"label":"All","value":"All"}]
    if cat and cat != "All":
        df_m = df_m[df_m["Category"]==cat]
    vendors = sorted([
        v for v in df_m["Vendor"].unique()
        if str(v).strip() not in ["","nan"]
    ])
    return ([{"label":"All","value":"All"}] +
            [{"label":v,"value":v} for v in vendors])


# ════════════════════════════════════════════════════════════
# RUN
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app.run(debug=True, port=8050)
